from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi import Header
from sqlalchemy.orm import Session, selectinload
from sqlalchemy import and_, or_
from app.db import get_db
from app.models.class_meeting import ClassMeeting
from app.models.subject import Subject
from app.models.user import User
from app.models.user_week_study_schedule import UserWeekStudySchedule
from pydantic import BaseModel, field_validator
from datetime import time
from typing import List, Optional, Tuple
import uuid
import base64
import csv
import re
import io
from openpyxl import load_workbook
router = APIRouter(prefix="/timetable", tags=["Timetable"])

TIME_HHMM_RE = re.compile(r"^(?:[01]\d|2[0-3]):[0-5]\d$")

# Pydantic Models


class ClassMeetingCreate(BaseModel):
    subject_id: str
    day_of_week: int  # 0-6 (0=Sun … 6=Sat)
    start_time: str  # "09:00"
    end_time: str
    rrule: Optional[str] = None

    @field_validator("day_of_week")
    @classmethod
    def validate_day_of_week(cls, v: int):
        if v < 0 or v > 6:
            raise ValueError("day_of_week must be between 0 and 6 (0=Sun … 6=Sat).")
        return v

    @field_validator("start_time", "end_time")
    @classmethod
    def validate_time_format(cls, v: str):
        # Enforce strict HH:MM (00:00–23:59). Rejects '1:00' and '24:00'.
        if not TIME_HHMM_RE.match(v):
            raise ValueError("Time must be in HH:MM format (00:00 to 23:59). Example: '09:00'.")
        return v


class ClassMeetingUpdate(BaseModel):
    day_of_week: Optional[int] = None
    start_time: Optional[str] = None
    end_time: Optional[str] = None
    rrule: Optional[str] = None
    # Convenience fields to update the linked subject
    subject_name: Optional[str] = None  # alias for Subject.title
    code: Optional[str] = None
    difficulty: Optional[int] = None
    target_grade: Optional[str] = None
    credit_weight: Optional[float] = None

    @field_validator("day_of_week")
    @classmethod
    def validate_day_of_week_optional(cls, v: Optional[int]):
        if v is None:
            return v
        if v < 0 or v > 6:
            raise ValueError("day_of_week must be between 0 and 6 (0=Sun … 6=Sat).")
        return v

    @field_validator("start_time", "end_time")
    @classmethod
    def validate_time_format_optional(cls, v: Optional[str]):
        if v is None:
            return v
        if not TIME_HHMM_RE.match(v):
            raise ValueError("Time must be in HH:MM format (00:00 to 23:59). Example: '09:00'.")
        return v


class SubjectCreate(BaseModel):
    """Create a subject for a user.

    NOTE: The original version of this router expected many extra fields
    (exam_date, location, importance, etc.). The actual SQLAlchemy Subject
    model only contains: title, code, difficulty, target_grade, credit_weight,
    and is_active.

    To remain backward compatible with earlier frontend code, we still accept
    `name` as an alias for `title`.
    """

    user_id: str
    # New canonical field
    title: Optional[str] = None
    # Backward-compatible alias
    name: Optional[str] = None
    code: Optional[str] = None
    difficulty: Optional[int] = None
    target_grade: Optional[str] = None
    credit_weight: Optional[float] = None

    @field_validator("title")
    @classmethod
    def _title_or_name_required(cls, v: Optional[str], info):
        # If title missing, try name; ensure at least one is present
        name = (info.data.get("name") or "").strip() if info.data else ""
        if (v is None or not str(v).strip()) and not name:
            raise ValueError("Either 'title' or 'name' is required")
        return v

    @field_validator("title")
    @classmethod
    def _title_non_empty(cls, v: Optional[str]) -> Optional[str]:
        if v is None:
            return v
        v = v.strip()
        if not v:
            raise ValueError("title must not be empty")
        return v

    def resolved_title(self) -> str:
        # prefer title; fall back to name
        t = (self.title or self.name or "").strip()
        if not t:
            raise ValueError("Either title or name must be provided")
        return t


# READ ENDPOINTS


# ===============================
# Image / CSV / Excel timetable extraction
# ===============================

from PIL import Image
import uuid

try:
    import pytesseract  # type: ignore
except Exception:  # pragma: no cover
    pytesseract = None

import os
import shutil


class CalendarSessionIn(BaseModel):
    id: Optional[str] = None
    subject: str
    day: int
    startTime: str
    endTime: str
    type: Optional[str] = None
    color: Optional[str] = None
    deadline: Optional[str] = None


class CalendarSessionOut(BaseModel):
    id: str
    subject: str
    startTime: str
    endTime: str
    day: int
    type: Optional[str] = None
    color: Optional[str] = None
    deadline: Optional[str] = None


@router.get("/user/{user_id}/sessions", response_model=List[CalendarSessionOut])
def get_user_calendar_sessions(
    user_id: str,
    week_id: str = Query("default"),
    db: Session = Depends(get_db),
):
    """Return the user's study sessions for a given week."""
    row = (
        db.query(UserWeekStudySchedule)
        .filter(UserWeekStudySchedule.user_id == user_id, UserWeekStudySchedule.week_id == week_id)
        .first()
    )
    stored = (row.sessions if row and isinstance(row.sessions, list) else [])

    out: List[CalendarSessionOut] = []
    for s in stored:
        if not isinstance(s, dict):
            continue
        try:
            out.append(
                CalendarSessionOut(
                    id=str(s.get("id") or uuid.uuid4()),
                    subject=str(s.get("subject") or ""),
                    startTime=str(s.get("startTime") or "08:00"),
                    endTime=str(s.get("endTime") or "09:00"),
                    day=int(s.get("day") if s.get("day") is not None else 0),
                    type=s.get("type"),
                    color=s.get("color"),
                    deadline=s.get("deadline"),
                )
            )
        except Exception:
            continue
    return out


@router.put("/user/{user_id}/sessions")
def put_user_calendar_sessions(
    user_id: str,
    payload: List[CalendarSessionIn],
    week_id: str = Query("default"),
    db: Session = Depends(get_db),
):
    """Replace the user's study sessions for a given week."""
    normalized = []
    for s in payload:
        if not isinstance(s, dict):
            continue
        normalized.append({
            "id": s.id or str(uuid.uuid4()),
            "subject": s.subject,
            "day": s.day,
            "startTime": s.startTime,
            "endTime": s.endTime,
            "type": s.type,
            "color": s.color,
            "deadline": s.deadline,
        })

    row = (
        db.query(UserWeekStudySchedule)
        .filter(UserWeekStudySchedule.user_id == user_id, UserWeekStudySchedule.week_id == week_id)
        .first()
    )
    if row:
        row.sessions = normalized
    else:
        row = UserWeekStudySchedule(user_id=user_id, week_id=week_id, sessions=normalized)
        db.add(row)
    db.commit()
    return {"ok": True, "week_id": week_id, "sessions": len(normalized)}


class TimetableExtractItem(BaseModel):
    day_of_week: Optional[int] = None  # 0-6 (Sun..Sat)
    day_label: Optional[str] = None
    start_time: Optional[str] = None  # HH:MM
    end_time: Optional[str] = None
    rrule: Optional[str] = None
    subject_title: Optional[str] = None
    subject_code: Optional[str] = None
    raw_line: Optional[str] = None

class TimetableImportItem(BaseModel):
    day_of_week: int
    start_time: str
    end_time: str
    subject_title: str
    subject_code: Optional[str] = None
    rrule: Optional[str] = None


class TimetableImportPreviewResponse(BaseModel):
    items: List[TimetableImportItem]
    warnings: List[str] = []
    errors: List[str] = []


class TimetableImportRequest(BaseModel):
    user_id: str
    items: List[TimetableExtractItem]
    create_missing_subjects: bool = True
    replace_existing: bool = False


class TimetableImportResponse(BaseModel):
    created_subjects: int
    reused_subjects: int
    created_meetings: int
    skipped_items: int
    warnings: List[str] = []
class TimetableExtractResponse(BaseModel):
    text: str
    items: List[TimetableExtractItem]


class TimetableCsvTextRequest(BaseModel):
    csv: str


class TimetableExtractBase64In(BaseModel):
    image_base64: str


_DAY_MAP = {
    "sun": 0, "sunday": 0, "s": 0,
    "mon": 1, "monday": 1, "m": 1,
    "tue": 2, "tues": 2, "tuesday": 2, "t": 2,
    "wed": 3, "wednesday": 3, "w": 3,
    "thu": 4, "thur": 4, "thurs": 4, "thursday": 4, "r": 4,
    "fri": 5, "friday": 5, "f": 5,
    "sat": 6, "saturday": 6, "sa": 6,
}

_TIME_RANGE_RE = re.compile(
    r"(?P<s>\b\d{1,2}(?::|\.)\d{2}\s*(?:am|pm)?\b)\s*[-–—]\s*(?P<e>\b\d{1,2}(?::|\.)\d{2}\s*(?:am|pm)?\b)",
    flags=re.IGNORECASE,
)

# stricter than before to reduce OCR garbage
_COURSE_CODE_RE = re.compile(r"^[A-Z]{2,6}\d{3,4}[A-Z]?$")
_SECTION_RE = re.compile(r"^(CR|ROOM|RM|LAB|SEC|SECTION)[A-Z0-9_-]*$", re.IGNORECASE)


def _ensure_tesseract_cmd() -> None:
    if pytesseract is None:
        return

    try:
        current = getattr(pytesseract.pytesseract, "tesseract_cmd", "") or ""
        if current and os.path.exists(current):
            return
    except Exception:
        pass

    exe = None
    try:
        exe = shutil.which("tesseract")
    except Exception:
        exe = None

    candidates = []
    if exe:
        candidates.append(exe)

    candidates.extend(
        [
            r"C:\\Program Files\\Tesseract-OCR\\tesseract.exe",
            r"C:\\Program Files (x86)\\Tesseract-OCR\\tesseract.exe",
        ]
    )

    for p in candidates:
        try:
            if p and os.path.exists(p):
                pytesseract.pytesseract.tesseract_cmd = p
                return
        except Exception:
            continue

    raise HTTPException(
        status_code=500,
        detail="Tesseract OCR is not installed or not on PATH. Install Tesseract and restart the backend.",
    )

def _is_meaningful_subject_title(value: Optional[str]) -> bool:
    if not value:
        return False

    s = re.sub(r"\s+", " ", str(value).strip())
    if not s:
        return False

    if _looks_like_section_or_room(s):
        return False

    if len(s) < 4:
        return False

    if re.fullmatch(r"[\d\W_]+", s):
        return False

    bad_words = {
        "room", "lab", "section", "sec", "rm", "cr",
        "time", "day", "course", "subject", "class", "lecture"
    }
    if s.lower() in bad_words:
        return False

    token = re.sub(r"[^A-Za-z0-9]", "", s).upper()
    if len(s.split()) == 1 and not _COURSE_CODE_RE.match(token):
        if not re.fullmatch(r"[A-Za-z]{3,}", s):
            return False

    return True


def _normalize_extract_item(it: TimetableExtractItem) -> Tuple[Optional[TimetableImportItem], Optional[str]]:
    if it.day_of_week is None:
        return None, f"Missing day_of_week for line: {it.raw_line or ''}"

    if not it.start_time or not it.end_time:
        return None, f"Missing time range for line: {it.raw_line or ''}"

    try:
        start_time = _norm_time(it.start_time)
        end_time = _norm_time(it.end_time)
    except Exception as e:
        return None, f"Invalid time in line '{it.raw_line or ''}': {e}"

    if start_time >= end_time:
        return None, f"start_time must be before end_time for line: {it.raw_line or ''}"

    subject_title = (it.subject_title or "").strip()
    subject_code = (it.subject_code or "").strip() or None

    if not _is_meaningful_subject_title(subject_title):
        if subject_code and _COURSE_CODE_RE.match(re.sub(r"[^A-Za-z0-9]", "", subject_code).upper()):
            subject_title = subject_code
        else:
            return None, f"Rejected weak/empty subject title for line: {it.raw_line or ''}"

    return TimetableImportItem(
        day_of_week=it.day_of_week,
        start_time=start_time,
        end_time=end_time,
        subject_title=subject_title,
        subject_code=subject_code,
        rrule=it.rrule,
    ), None


def _normalize_extract_items(items: List[TimetableExtractItem]) -> Tuple[List[TimetableImportItem], List[str]]:
    normalized: List[TimetableImportItem] = []
    errors: List[str] = []

    for it in items:
        norm, err = _normalize_extract_item(it)
        if err:
            errors.append(err)
            continue
        if norm:
            normalized.append(norm)

    # de-duplicate normalized items
    seen = set()
    uniq: List[TimetableImportItem] = []
    for it in normalized:
        key = (it.day_of_week, it.start_time, it.end_time, it.subject_code or it.subject_title)
        if key in seen:
            continue
        seen.add(key)
        uniq.append(it)

    return uniq, errors

def _norm_time(t: str) -> str:
    s = str(t or "").strip().lower().replace(".", ":")
    s = re.sub(r"\s+", "", s)

    m = re.match(r"^(\d{1,2}):(\d{2})(am|pm)?$", s)
    if not m:
        raise ValueError(f"Invalid time format: {t}")

    hh = int(m.group(1))
    mm = int(m.group(2))
    ampm = m.group(3)

    if mm < 0 or mm > 59:
        raise ValueError(f"Invalid minutes in time: {t}")

    if ampm:
        if hh < 1 or hh > 12:
            raise ValueError(f"Invalid 12-hour time: {t}")
        if ampm == "am":
            hh = 0 if hh == 12 else hh
        else:
            hh = 12 if hh == 12 else hh + 12
    else:
        if hh < 0 or hh > 23:
            raise ValueError(f"Invalid 24-hour time: {t}")

    return f"{hh:02d}:{mm:02d}"

def _find_or_create_subject(
    db: Session,
    user_id: str,
    subject_title: str,
    subject_code: Optional[str],
    create_missing: bool = True,
) -> Tuple[Optional[Subject], bool]:
    query = db.query(Subject).filter(Subject.user_id == user_id)

    if subject_code:
        subject = query.filter(Subject.code == subject_code).first()
        if subject:
            return subject, False

    subject = query.filter(Subject.title == subject_title).first()
    if subject:
        return subject, False

    if not create_missing:
        return None, False

    subject = Subject(
        id=str(uuid.uuid4()),
        user_id=user_id,
        title=subject_title,
        code=subject_code,
        is_active=True,
    )
    db.add(subject)
    db.flush()
    return subject, True
def _meeting_exists(
    db: Session,
    subject_id: str,
    day_of_week: int,
    start_time: str,
    end_time: str,
) -> bool:
    existing = (
        db.query(ClassMeeting)
        .filter(
            ClassMeeting.subject_id == subject_id,
            ClassMeeting.day_of_week == day_of_week,
            ClassMeeting.start_time == start_time,
            ClassMeeting.end_time == end_time,
        )
        .first()
    )
    return existing is not None
def _day_to_int(val: str | int | None) -> Optional[int]:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if s.isdigit():
        d = int(s)
        return d if 0 <= d <= 6 else None
    key = re.sub(r"[^a-z]", "", s.lower())
    if key in _DAY_MAP:
        return _DAY_MAP[key]
    if len(key) >= 3 and key[:3] in _DAY_MAP:
        return _DAY_MAP[key[:3]]
    return None


def _infer_day(line: str) -> Tuple[Optional[int], Optional[str]]:
    lower = re.sub(r"[^a-z]", " ", line.lower())
    tokens = [t for t in lower.split() if t]
    for tok in tokens:
        if tok in _DAY_MAP:
            return _DAY_MAP[tok], tok
    return None, None


def _looks_like_section_or_room(value: str | None) -> bool:
    if not value:
        return False
    raw = re.sub(r"[^A-Za-z0-9_-]", "", value).upper()
    return bool(_SECTION_RE.match(raw))


def _clean_subject_value(title: Optional[str], code: Optional[str]) -> Tuple[Optional[str], Optional[str]]:
    title = (title or "").strip() or None
    code = (code or "").strip() or None

    if title and _looks_like_section_or_room(title):
        title = None
    if code and _looks_like_section_or_room(code):
        code = None

    if not title and code and _COURSE_CODE_RE.match(code):
        title = code

    return title, code


def _infer_subject(line: str, time_span: Optional[Tuple[str, str]]) -> Tuple[Optional[str], Optional[str]]:
    cleaned = line

    if time_span:
        s, e = time_span
        cleaned = cleaned.replace(s, " ").replace(e, " ")

    for k in _DAY_MAP.keys():
        cleaned = re.sub(rf"\b{k}\b", " ", cleaned, flags=re.IGNORECASE)

    cleaned = re.sub(r"\bsection\s*:\s*[A-Za-z0-9_-]+\b", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\broom\s*:\s*[A-Za-z0-9_-]+\b", " ", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" -|\t")

    m = re.search(r"\b([A-Z]{2,6}\s?[-]?\s?\d{3,4}[A-Z]?)\b", cleaned)
    code = None
    title = None

    if m:
        code = re.sub(r"\s+", "", m.group(1)).replace("-", "")
        title = code
    else:
        title = cleaned

    title, code = _clean_subject_value(title, code)
    return title, code


def _is_complete_item(it: TimetableExtractItem) -> bool:
    return (
        it.day_of_week is not None
        and bool(it.start_time)
        and bool(it.end_time)
        and bool(it.subject_title)
    )


def _dedupe_items(items: List[TimetableExtractItem]) -> List[TimetableExtractItem]:
    seen = set()
    out: List[TimetableExtractItem] = []
    for it in items:
        key = (it.day_of_week, it.start_time, it.end_time, it.subject_code or it.subject_title)
        if key in seen:
            continue
        seen.add(key)
        out.append(it)
    return out


def _score_items(items: List[TimetableExtractItem]) -> int:
    score = 0
    seen = set()

    for it in items:
        if _is_complete_item(it):
            score += 10
        if it.subject_code and _COURSE_CODE_RE.match(it.subject_code):
            score += 4
        if it.subject_title and not _looks_like_section_or_room(it.subject_title):
            score += 2

        key = (it.day_of_week, it.start_time, it.end_time, it.subject_title, it.subject_code)
        if key in seen:
            score -= 3
        seen.add(key)

    return score


def _extract_items_line_based(text: str) -> List[TimetableExtractItem]:
    lines = [ln.strip() for ln in (text or "").splitlines() if ln.strip()]
    items: List[TimetableExtractItem] = []

    for ln in lines:
        day_idx, day_label = _infer_day(ln)
        m = _TIME_RANGE_RE.search(ln)

        time_span = None
        start_time = end_time = None
        if m:
            try:
                start_time = _norm_time(m.group("s"))
                end_time = _norm_time(m.group("e"))
                time_span = (m.group("s"), m.group("e"))
            except Exception:
                start_time = end_time = None

        title, code = _infer_subject(ln, time_span)

        if start_time or end_time or day_idx is not None:
            items.append(
                TimetableExtractItem(
                    day_of_week=day_idx,
                    day_label=day_label,
                    start_time=start_time,
                    end_time=end_time,
                    subject_title=title,
                    subject_code=code,
                    raw_line=ln,
                )
            )
    return items


def _is_valid_course_token(
    raw: str,
    conf: float,
    left: int,
    top: int,
    width: int,
    height: int,
    header_y: float,
    left_x: float,
) -> bool:
    token = re.sub(r"[^A-Za-z0-9]", "", (raw or "")).upper()
    if not token:
        return False
    if conf < 55:
        return False
    if len(token) < 5 or len(token) > 12:
        return False
    if not _COURSE_CODE_RE.match(token):
        return False
    if _looks_like_section_or_room(token):
        return False
    if top <= header_y:
        return False
    if left <= left_x:
        return False
    if width < 18 or height < 10:
        return False
    return True


def _extract_items_grid(img: Image.Image) -> List[TimetableExtractItem]:
    if pytesseract is None:
        return []

    ocr_img = img
    try:
        w0, h0 = img.size
        scale = 1
        if max(w0, h0) < 1400:
            scale = 2
        if max(w0, h0) < 900:
            scale = 3
        if scale > 1:
            ocr_img = img.resize((w0 * scale, h0 * scale), Image.Resampling.LANCZOS)

        from PIL import ImageEnhance, ImageOps
        gray = ImageOps.grayscale(ocr_img)
        gray = ImageEnhance.Contrast(gray).enhance(1.6)
        gray = ImageEnhance.Sharpness(gray).enhance(1.8)
        ocr_img = gray.convert("RGB")
    except Exception:
        ocr_img = img

    try:
        from pytesseract import Output  # type: ignore
        data = pytesseract.image_to_data(ocr_img, output_type=Output.DICT, config="--psm 4")
    except Exception:
        try:
            from pytesseract import Output  # type: ignore
            data = pytesseract.image_to_data(ocr_img, output_type=Output.DICT)
        except Exception:
            return []

    W, H = ocr_img.size
    n = len(data.get("text", []))
    words = []

    for i in range(n):
        t = (data.get("text", [""])[i] or "").strip()
        if not t:
            continue

        try:
            conf = float(str(data.get("conf", ["-1"])[i]))
        except Exception:
            conf = -1.0
        if conf < 40:
            continue

        try:
            left = int(data.get("left", [0])[i])
            top = int(data.get("top", [0])[i])
            w = int(data.get("width", [0])[i])
            h = int(data.get("height", [0])[i])
        except Exception:
            continue

        words.append(
            {
                "text": t,
                "conf": conf,
                "left": left,
                "top": top,
                "width": w,
                "height": h,
                "x": left + (w / 2.0),
                "y": top + (h / 2.0),
                "block_num": data.get("block_num", [None])[i],
                "par_num": data.get("par_num", [None])[i],
                "line_num": data.get("line_num", [None])[i],
            }
        )

    if not words:
        return []

    header_y = H * 0.18
    left_x = W * 0.25

    day_hits = {}
    for w in words:
        if w["top"] > header_y:
            continue
        key = re.sub(r"[^a-z]", "", w["text"].lower())
        if len(key) >= 3 and key[:3] in _DAY_MAP:
            dk = key if key in _DAY_MAP else key[:3]
            di = _DAY_MAP[dk]
            day_hits.setdefault(di, []).append(w["x"])

    day_cols = []
    for di, xs in day_hits.items():
        day_cols.append((sum(xs) / max(len(xs), 1), di))
    day_cols.sort(key=lambda x: x[0])

    if len(day_cols) < 2:
        margin = W * 0.15
        labels = [1, 2, 3, 4, 5]
        col_w = (W - margin) / len(labels)
        day_cols = [(margin + col_w * (i + 0.5), labels[i]) for i in range(len(labels))]

    xs = [x for x, _ in day_cols]
    bounds = [0.0]
    for i in range(len(xs) - 1):
        bounds.append((xs[i] + xs[i + 1]) / 2.0)
    bounds.append(float(W))

    def day_for_x(x: float) -> Optional[int]:
        for i in range(len(xs)):
            if bounds[i] <= x < bounds[i + 1]:
                return day_cols[i][1]
        return None

    left_words = [w for w in words if w["left"] < left_x and w["top"] > header_y]
    line_groups = {}
    for w in left_words:
        k = (w.get("block_num"), w.get("par_num"), w.get("line_num"))
        line_groups.setdefault(k, []).append(w)

    time_rows = []
    for _, ws in line_groups.items():
        ws = sorted(ws, key=lambda z: z["left"])
        line_text = " ".join([z["text"] for z in ws])
        m = re.search(r"\b(\d{1,2})\s*(AM|PM)?\b", line_text, flags=re.I)
        if not m:
            continue
        hour = int(m.group(1))
        ampm = (m.group(2) or "").upper()
        y = sum([z["y"] for z in ws]) / max(len(ws), 1)
        time_rows.append((y, hour, ampm))

    time_rows.sort(key=lambda t: t[0])
    if not time_rows:
        return []

    start_times = []
    prev_hour = None
    is_pm = False
    for y, hour, ampm in time_rows:
        if ampm == "PM":
            is_pm = True
        elif ampm == "AM":
            is_pm = False
        else:
            if prev_hour is not None and hour < prev_hour:
                is_pm = True
        prev_hour = hour

        if ampm == "AM":
            hh = 0 if hour == 12 else hour
        elif ampm == "PM":
            hh = 12 if hour == 12 else hour + 12
        else:
            hh = 12 if (is_pm and hour == 12) else (hour + 12 if is_pm and hour != 12 else (0 if hour == 12 else hour))

        start_times.append((y, hh))

    diffs = [start_times[i + 1][0] - start_times[i][0] for i in range(len(start_times) - 1)]
    diffs = [d for d in diffs if d > 0]
    diffs.sort()
    row_h = diffs[len(diffs) // 2] if diffs else (H * 0.08)

    def nearest_start(y: float) -> Optional[int]:
        if not start_times:
            return None
        yy, hh = min(start_times, key=lambda t: abs(t[0] - y))
        if abs(yy - y) > row_h * 0.9:
            return None
        return hh

    items: List[TimetableExtractItem] = []
    for w in words:
        raw = re.sub(r"[^A-Za-z0-9]", "", w["text"]).upper()

        if not _is_valid_course_token(
            raw=raw,
            conf=float(w.get("conf", 0)),
            left=int(w["left"]),
            top=int(w["top"]),
            width=int(w["width"]),
            height=int(w["height"]),
            header_y=header_y,
            left_x=left_x,
        ):
            continue

        day_idx = day_for_x(w["x"])
        if day_idx is None:
            continue

        hh = nearest_start(w["y"])
        if hh is None:
            continue

        start_time = f"{hh:02d}:00"

        slots = 1
        try:
            slots = max(1, int(round(w["height"] / max(row_h, 1))))
        except Exception:
            slots = 1

        end_total = (hh * 60) + (slots * 50)
        end_h = end_total // 60
        end_m = end_total % 60
        end_time = f"{end_h:02d}:{end_m:02d}"

        subject_title, subject_code = _clean_subject_value(raw, raw)
        if not subject_title and not subject_code:
            continue

        items.append(
            TimetableExtractItem(
                day_of_week=day_idx,
                start_time=start_time,
                end_time=end_time,
                subject_title=subject_title,
                subject_code=subject_code,
                raw_line=w["text"],
            )
        )

    uniq = _dedupe_items(items)

    filtered: List[TimetableExtractItem] = []
    by_title: dict[str, int] = {}
    for it in uniq:
        title = (it.subject_title or "").strip().upper()
        by_title[title] = by_title.get(title, 0) + 1
        if by_title[title] > 4:
            continue
        filtered.append(it)

    if len(filtered) > 8:
        filtered = filtered[:8]

    return filtered


def _read_csv_text(raw: str) -> TimetableExtractResponse:
    f = io.StringIO(raw)
    reader = csv.DictReader(f)
    if not reader.fieldnames:
        raise HTTPException(status_code=400, detail="CSV has no headers/columns")

    def norm(h: str) -> str:
        return re.sub(r"[^a-z0-9]", "", h.strip().lower())

    headers = {norm(h): h for h in reader.fieldnames}

    def get(row, *names):
        for n in names:
            k = norm(n)
            if k in headers:
                return row.get(headers[k])
        return None

    items: List[TimetableExtractItem] = []
    for row in reader:
        day = _day_to_int(get(row, "day", "day_of_week"))
        start_raw = get(row, "start", "start_time")
        end_raw = get(row, "end", "end_time")
        if not start_raw or not end_raw:
            continue

        try:
            start = _norm_time(str(start_raw))
            end = _norm_time(str(end_raw))
        except Exception:
            continue

        code = get(row, "code", "subject_code")
        title = get(row, "subject", "title", "course", "course_title")
        subject_title, subject_code = _clean_subject_value(
            str(title).strip() if title else None,
            str(code).strip() if code else None,
        )

        items.append(
            TimetableExtractItem(
                day_of_week=day,
                start_time=start,
                end_time=end,
                subject_title=subject_title,
                subject_code=subject_code,
                raw_line=",".join([str(row.get(h, "")).strip() for h in reader.fieldnames]),
            )
        )

    return TimetableExtractResponse(text=raw, items=_dedupe_items(items))


def _read_excel_bytes(raw_bytes: bytes) -> TimetableExtractResponse:
    try:
        wb = load_workbook(filename=io.BytesIO(raw_bytes), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {e}")

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Excel file is empty")

    headers = [str(v).strip() if v is not None else "" for v in rows[0]]
    if not any(headers):
        raise HTTPException(status_code=400, detail="Excel file has no header row")

    def norm(h: str) -> str:
        return re.sub(r"[^a-z0-9]", "", str(h).strip().lower())

    header_map = {norm(h): idx for idx, h in enumerate(headers) if h}

    def get(row, *names):
        for n in names:
            idx = header_map.get(norm(n))
            if idx is not None and idx < len(row):
                return row[idx]
        return None

    items: List[TimetableExtractItem] = []

    for row in rows[1:]:
        day = _day_to_int(get(row, "day", "day_of_week"))
        start_raw = get(row, "start", "start_time")
        end_raw = get(row, "end", "end_time")
        title = get(row, "subject", "title", "course", "course_title")
        code = get(row, "code", "subject_code")

        if not start_raw or not end_raw:
            continue

        try:
            start = _norm_time(str(start_raw))
            end = _norm_time(str(end_raw))
        except Exception:
            continue

        subject_title, subject_code = _clean_subject_value(
            str(title).strip() if title else None,
            str(code).strip() if code else None,
        )

        items.append(
            TimetableExtractItem(
                day_of_week=day,
                start_time=start,
                end_time=end,
                subject_title=subject_title,
                subject_code=subject_code,
                raw_line=" | ".join("" if v is None else str(v) for v in row),
            )
        )

    raw_text = "\n".join(", ".join("" if v is None else str(v) for v in row) for row in rows)
    return TimetableExtractResponse(text=raw_text, items=_dedupe_items(items))


@router.post("/extract-image", response_model=TimetableExtractResponse)
async def extract_timetable_from_image(file: UploadFile = File(...)):
    if pytesseract is None:
        raise HTTPException(
            status_code=500,
            detail="pytesseract is not installed on the server. Install Tesseract + pytesseract.",
        )

    _ensure_tesseract_cmd()

    if not file.content_type or not file.content_type.startswith("image/"):
        raise HTTPException(status_code=400, detail="Please upload an image file.")

    data = await file.read()
    try:
        img = Image.open(io.BytesIO(data)).convert("RGB")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid image: {e}")

    try:
        text = pytesseract.image_to_string(img, config="--psm 6")
    except Exception:
        text = pytesseract.image_to_string(img)

    # For timetable screenshots, prefer strict grid parsing only.
    items = _dedupe_items(_extract_items_grid(img))
    return TimetableExtractResponse(text=text, items=items)


@router.post("/extract-image-base64", response_model=TimetableExtractResponse)
def extract_timetable_from_image_base64(payload: TimetableExtractBase64In):
    if pytesseract is None:
        raise HTTPException(
            status_code=500,
            detail="pytesseract is not installed on the server. Install Tesseract + pytesseract.",
        )

    _ensure_tesseract_cmd()

    b64 = payload.image_base64.strip()
    if b64.lower().startswith("data:") and "," in b64:
        b64 = b64.split(",", 1)[1]

    try:
        raw = base64.b64decode(b64, validate=True)
    except Exception:
        try:
            raw = base64.b64decode(re.sub(r"\s+", "", b64))
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid base64 image: {e}")

    try:
        img = Image.open(io.BytesIO(raw)).convert("RGB")
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid image: {e}")

    try:
        text = pytesseract.image_to_string(img, config="--psm 6")
    except Exception:
        text = pytesseract.image_to_string(img)

    items = _dedupe_items(_extract_items_grid(img))
    return TimetableExtractResponse(text=text, items=items)


@router.post("/extract-csv", response_model=TimetableExtractResponse)
async def extract_timetable_from_csv(file: UploadFile = File(...)):
    if not file.filename.lower().endswith(".csv"):
        if file.content_type not in ("text/csv", "application/csv", "application/vnd.ms-excel"):
            raise HTTPException(status_code=400, detail="Please upload a .csv file")

    raw_bytes = await file.read()
    try:
        raw = raw_bytes.decode("utf-8")
    except UnicodeDecodeError:
        raw = raw_bytes.decode("utf-8-sig", errors="replace")

    return _read_csv_text(raw)


@router.post("/extract-csv-text", response_model=TimetableExtractResponse)
async def extract_timetable_from_csv_text(payload: TimetableCsvTextRequest):
    return _read_csv_text(payload.csv)


@router.post("/extract-excel", response_model=TimetableExtractResponse)
async def extract_timetable_from_excel(file: UploadFile = File(...)):
    filename = (file.filename or "").lower()
    if not (filename.endswith(".xlsx") or filename.endswith(".xlsm")):
        raise HTTPException(status_code=400, detail="Please upload a .xlsx or .xlsm file")

    raw_bytes = await file.read()
    return _read_excel_bytes(raw_bytes)


@router.post("/extract-file", response_model=TimetableExtractResponse)
async def extract_timetable_from_file(file: UploadFile = File(...)):
    filename = (file.filename or "").lower()
    content_type = file.content_type or ""
    raw_bytes = await file.read()

    if content_type.startswith("image/") or filename.endswith((".png", ".jpg", ".jpeg", ".webp")):
        if pytesseract is None:
            raise HTTPException(status_code=500, detail="pytesseract is not installed on the server.")
        _ensure_tesseract_cmd()

        try:
            img = Image.open(io.BytesIO(raw_bytes)).convert("RGB")
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Invalid image: {e}")

        try:
            text = pytesseract.image_to_string(img, config="--psm 6")
        except Exception:
            text = pytesseract.image_to_string(img)

        items = _dedupe_items(_extract_items_grid(img))
        return TimetableExtractResponse(text=text, items=items)

    if filename.endswith(".csv") or content_type in ("text/csv", "application/csv", "application/vnd.ms-excel"):
        try:
            raw = raw_bytes.decode("utf-8")
        except UnicodeDecodeError:
            raw = raw_bytes.decode("utf-8-sig", errors="replace")
        return _read_csv_text(raw)

    if filename.endswith(".xlsx") or filename.endswith(".xlsm"):
        return _read_excel_bytes(raw_bytes)

    raise HTTPException(status_code=400, detail="Unsupported file type")
@router.post("/import-preview", response_model=TimetableImportPreviewResponse)
async def import_timetable_preview(file: UploadFile = File(...)):
    extracted = await extract_timetable_from_file(file)
    normalized, errors = _normalize_extract_items(extracted.items)

    return TimetableImportPreviewResponse(
    items=normalized,
    warnings=[],
    errors=errors,
    )
@router.post("/import-items", response_model=TimetableImportResponse)
async def import_timetable_items(
    payload: TimetableImportRequest,
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == payload.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    normalized, errors = _normalize_extract_items(payload.items)

    created_subjects = 0
    reused_subjects = 0
    created_meetings = 0
    skipped_items = 0
    warnings = list(errors)

    if payload.replace_existing:
        subject_ids = [
            s.id for s in db.query(Subject.id).filter(Subject.user_id == payload.user_id).all()
        ]
        if subject_ids:
            db.query(ClassMeeting).filter(ClassMeeting.subject_id.in_(subject_ids)).delete(synchronize_session=False)

    for item in normalized:
        subject, created = _find_or_create_subject(
            db=db,
            user_id=payload.user_id,
            subject_title=item.subject_title,
            subject_code=item.subject_code,
            create_missing=payload.create_missing_subjects,
        )

        if not subject:
            skipped_items += 1
            warnings.append(f"Skipped '{item.subject_title}' because subject does not exist")
            continue

        if created:
            created_subjects += 1
        else:
            reused_subjects += 1

        if _meeting_exists(
            db=db,
            subject_id=subject.id,
            day_of_week=item.day_of_week,
            start_time=item.start_time,
            end_time=item.end_time,
        ):
            skipped_items += 1
            warnings.append(
                f"Duplicate skipped: {item.subject_title} day={item.day_of_week} {item.start_time}-{item.end_time}"
            )
            continue

        meeting = ClassMeeting(
            id=str(uuid.uuid4()),
            subject_id=subject.id,
            day_of_week=item.day_of_week,
            start_time=item.start_time,
            end_time=item.end_time,
            rrule=item.rrule,
        )
        db.add(meeting)
        created_meetings += 1

    db.commit()

    return TimetableImportResponse(
        created_subjects=created_subjects,
        reused_subjects=reused_subjects,
        created_meetings=created_meetings,
        skipped_items=skipped_items,
        warnings=warnings,
    )
@router.post("/import-items", response_model=TimetableImportResponse)
async def import_timetable_items(
    payload: TimetableImportRequest,
    db: Session = Depends(get_db),
):
    user = db.query(User).filter(User.id == payload.user_id).first()
    if not user:
        raise HTTPException(status_code=404, detail="User not found")

    normalized, errors = _normalize_extract_items(payload.items)

    created_subjects = 0
    reused_subjects = 0
    created_meetings = 0
    skipped_items = 0
    warnings = list(errors)

    if payload.replace_existing:
        subject_ids = [
            s.id for s in db.query(Subject.id).filter(Subject.user_id == payload.user_id).all()
        ]
        if subject_ids:
            db.query(ClassMeeting).filter(ClassMeeting.subject_id.in_(subject_ids)).delete(synchronize_session=False)

    for item in normalized:
        subject, created = _find_or_create_subject(
            db=db,
            user_id=payload.user_id,
            subject_title=item.subject_title,
            subject_code=item.subject_code,
            create_missing=payload.create_missing_subjects,
        )

        if not subject:
            skipped_items += 1
            warnings.append(f"Skipped '{item.subject_title}' because subject does not exist")
            continue

        if created:
            created_subjects += 1
        else:
            reused_subjects += 1

        if _meeting_exists(
            db=db,
            subject_id=subject.id,
            day_of_week=item.day_of_week,
            start_time=item.start_time,
            end_time=item.end_time,
        ):
            skipped_items += 1
            warnings.append(
                f"Duplicate skipped: {item.subject_title} day={item.day_of_week} {item.start_time}-{item.end_time}"
            )
            continue

        meeting = ClassMeeting(
            id=str(uuid.uuid4()),
            subject_id=subject.id,
            day_of_week=item.day_of_week,
            start_time=item.start_time,
            end_time=item.end_time,
            rrule=item.rrule,
        )
        db.add(meeting)
        created_meetings += 1

    db.commit()

    return TimetableImportResponse(
        created_subjects=created_subjects,
        reused_subjects=reused_subjects,
        created_meetings=created_meetings,
        skipped_items=skipped_items,
        warnings=warnings,
    )